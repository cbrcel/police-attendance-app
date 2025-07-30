from flask import Flask, render_template, request
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

app = Flask(__name__)

@app.route("/", methods=["GET", "POST"])
def index():
    status = ""
    police_stations = [
        "SDPO PALAKKAD", "TOWN SOUTH", "CONTROL ROOM", "TRAFFIC PALAKKAD", "TOWN NORTH", "MANKARA", "KASABA", "WALAYAR",
        "HEMAMBIKA NAGAR", "MALAMPUZHA", "SDPO CHITTUR", "CHITTUR", "MEENAKSHIPURAM", "KOZHINJAMPARA", "KOLLENGODE",
        "PUDUNAGARAM", "PARAMBIKULAM", "SDPO ALATHUR", "ALATHUR", "KOTTAYI", "KUZHALMANNAM", "VADAKKENCHERRY",
        "MANGALAMDAM", "NEMMARA", "PADAGIRI", "SDPO SHORNUR", "OTTAPALAM", "OTTAPALAM TRAFFIC", "SHORNUR", "PATTAMBI",
        "PATTAMBI TRAFFIC", "KOPPAM", "THRITHALA", "CHALISSERY", "SDPO MANNARKKAD", "CHERPULASSERY", "SREEKRISHNAPURAM",
        "MANNARKKAD", "NATTUKAL", "KONGAD", "MANNARKKAD TRAFFIC", "KALLADIKKODE", "SMS AGALI", "AGALI", "SHOLAYUR",
        "PUDUR", "SPECIAL BRANCH", "DCRB", "DCB", "CYBER CELL", "CYBER CRIME", "NARCOTIC CELL", "WOMEN CELL",
        "VANITHA", "TOURISM POLICE", "DCPHQ", "DHQ"
    ]

    ranks = [
        "DySP", "IoP", "SI", "ASI", "GSI", "GASI", "SCPO", "GSCPO", "CPO", "WSCPO", "WCPO", "AC", "RSI", "RASI",
        "MTSI", "GRSI", "GRASI", "Armr SI", "Armr GASI", "Armr SCPO", "Armr CPO", "KHG", "RTPC", "CF"
    ]

    if request.method == "POST":
        officer = request.form.get("officer")
        rank = request.form.get("rank")
        station = request.form.get("station")
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        df = pd.DataFrame([[officer, rank, station, now]], columns=["Officer", "Rank", "Police Station", "Timestamp"])
        excel_path = "attendance.xlsx"
        try:
            book = load_workbook(excel_path)
            writer = pd.ExcelWriter(excel_path, engine="openpyxl")
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            start_row = writer.sheets["Sheet1"].max_row
            df.to_excel(writer, index=False, header=False, startrow=start_row)
            writer.close()
        except FileNotFoundError:
            df.to_excel(excel_path, index=False)
        status = "Attendance recorded successfully."

    return render_template("index.html", status=status, stations=police_stations, ranks=ranks)

if __name__ == "__main__":
    app.run(debug=True)
